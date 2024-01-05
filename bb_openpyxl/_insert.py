# -*- coding: utf-8 -*-
"""
@author: hubo
@project: bb-py
@file: _insert.py.py
@time: 2023/12/22 14:58
@desc:
"""
import copy
import warnings
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from ._range import _new_range
from ._formula import _reset_formula


def _get_all_columns_width(ws):
    """在插入或者删除列之前获取列宽"""
    widths = []
    for c in range(1, ws.max_column):
        widths.append(ws.column_dimensions[get_column_letter(c)].width)
    return widths


def _reset_all_columns_width(ws, widths, idx, amount=1):
    """在插入或者删除列之后重新设置列宽"""
    if amount > 0:  # 插入列
        for c in range(idx+amount, ws.max_column):
            ws.column_dimensions[get_column_letter(c)].width = widths[c - amount - 1]
    else:   # 删除列
        for c in range(idx, ws.max_column):
            ws.column_dimensions[get_column_letter(c)].width = widths[c - amount - 1]


def _get_all_rows_height(ws):
    """在插入或者删除行之前获取行高"""
    heights = []
    for r in range(1, ws.max_row):
        heights.append(ws.row_dimensions[r].height)
    return heights


def _reset_all_rows_height(ws, heights, idx, amount=1):
    """在插入或者删除列之后重新设置列宽"""
    if amount > 0:
        for r in range(idx+amount, ws.max_row):
            ws.row_dimensions[r].height = heights[r - amount - 1]
    else:
        for r in range(idx, ws.max_column):
            ws.row_dimensions[r].height = heights[r - amount - 1]


def _un_merge_cells_before_insert(ws, row_idx=None, col_idx=None, amount=1):
    """un_merge操作必须在插入行或者列之前"""
    unmerged_ranges = []
    for merged_range in ws.merged_cells.ranges:
        start_column, start_row, end_column, end_row = merged_range.bounds
        if (col_idx is not None and (col_idx <= start_column or col_idx <= end_column))\
                or (row_idx is not None and (row_idx <= start_row or row_idx <= end_row)):
            unmerged_ranges.append(copy.copy(merged_range))
    for merged_range in unmerged_ranges:
        ws.unmerge_cells(merged_range.coord)
    return unmerged_ranges


def _re_merge_cells_when_after_insert(ws, unmerged_ranges, row_idx=None, col_idx=None, amount=1):
    """在插入列时，重新计算合并的单元格"""
    for merged_range in unmerged_ranges:
        start_column, start_row, end_column, end_row = merged_range.bounds
        new_start_column, new_start_row, new_end_column, new_end_row = _new_range(start_column, start_row, end_column,
                                                                                  end_row, row_idx, col_idx, amount)
        # 判断是否还有merge的必要
        if ((new_end_column > new_start_column or new_end_row > new_start_row)
                and new_end_column >= new_start_column and new_end_row >= new_start_row):
            new_merged_range = f"{ws.cell(row=new_start_row, column=new_start_column).coordinate}:{ws.cell(row=new_end_row, column=new_end_column).coordinate}"
            ws.merge_cells(new_merged_range)


def _get_sheet_title_from_defined_name(text):
    if text.startswith('\''):
        next_pos = text.find('\'', 1)
        return text[1:next_pos]
    next_pos = text.find('!')
    return text[:next_pos]


def _re_set_all_formulas(ws, row_idx=None, col_idx=None, amount=1):
    """重设公式，在做了插入或者删除行列之后"""
    # 遍历每一行
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        # 遍历每个单元格
        for cell in row:
            if cell.data_type == 'f':
                cell.value = '=' + _reset_formula(cell.value[1:], row_idx=row_idx, col_idx=col_idx, amount=amount)

    # 处理命名范围
    for k, v in ws.parent.defined_names.items():
        content = v.attr_text
        if _get_sheet_title_from_defined_name(content) == ws.title:
            v.attr_text = _reset_formula(content, row_idx=row_idx, col_idx=col_idx, amount=amount)


def _warning_unsupported_formula(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.data_type == 'f':
                    if cell.value.find('!') >= 0:
                        warnings.warn('文件存在跨sheet页引用数据的公式。增删行或者列的操作，可能导致公式失效。建议通过定义名称规避此问题')
                        break


def _re_set_tables(ws, row_idx=None, col_idx=None, amount=1):
    tables = []
    for table in ws.tables.values():
        tab = Table(displayName=table.displayName)
        for attribute_name in dir(table):
            if (attribute_name in
                    ('ref', 'name', 'comment', 'tableType', 'headerRowCount', 'insertRow',
                     'insertRowShift', 'totalsRowCount', 'totalsRowShown', 'published', 'headerRowDxfId',
                     'dataDxfId', 'totalsRowDxfId', 'headerRowBorderDxfId', 'tableBorderDxfId',
                     'totalsRowBorderDxfId', 'headerRowCellStyle', 'dataCellStyle', 'totalsRowCellStyle',
                     'connectionId', 'autoFilter', 'sortState', 'tableStyleInfo')):
                setattr(tab, attribute_name, copy.copy(getattr(table, attribute_name)))
        tables.append(tab)
    ws.tables.clear()
    for table in tables:
        new_ref = _reset_formula(table.ref, row_idx=row_idx, col_idx=col_idx, amount=amount)
        # 如果新增了列，要增加新标题，否则会报错
        if new_ref != table.ref:
            table.ref = new_ref
            for i in range(0, table.headerRowCount):
                ...

        ws.add_table(table)
